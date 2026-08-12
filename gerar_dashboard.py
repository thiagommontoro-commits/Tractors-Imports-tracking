#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
==============================================================================
 Comex Stat API - Importação de TRATORES por país, ano/mês e segmento (HP)
 -> SEMPRE traz o mais recente: de 2022-01 até o último mês publicado.
 -> Gera Excel (padrão AGCO) e um index.html autossuficiente (arquivo único).
==============================================================================
Fonte: Ministério do Desenvolvimento, Indústria, Comércio e Serviços (MDIC)
Autor do script: Global Reporting & Analytics - AGCO (Thiago Montoro)
Dependências: requests, pandas, openpyxl  ->  pip install requests pandas openpyxl
==============================================================================
"""
import os, re, time, json, requests
import pandas as pd
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "https://api-comexstat.mdic.gov.br/general"
DATES_URL = "https://api-comexstat.mdic.gov.br/general/dates/updated"
ANO_INICIO = "2022-01"
PASTA_DESTINO = r"C:\Users\tm75667\OneDrive - AGCO Corp\Área de Trabalho\AI Projects\ComexStat"
NCMS_TRATORES = ["87019100","87019200","87019300","87019410","87019490","87019510","87019590"]
PREFIXO_SEGMENTO = {"870191":("Até 24 HP","Micro/compactos"),"870192":("Acima de 24 a 50 HP","Pequenos"),
    "870193":("Acima de 50 a 100 HP","Médios"),"870194":("Acima de 101 a 175 HP","Grandes"),
    "870195":("Acima de 175 HP","Alta Potência")}
ORDEM_HP = ["Até 24 HP","Acima de 24 a 50 HP","Acima de 50 a 100 HP","Acima de 101 a 175 HP","Acima de 175 HP"]
MESES_PT = {1:"Jan",2:"Fev",3:"Mar",4:"Abr",5:"Mai",6:"Jun",7:"Jul",8:"Ago",9:"Set",10:"Out",11:"Nov",12:"Dez"}
AGCO_VERMELHO="C8102E"; AGCO_CINZA="53565A"; BRANCO="FFFFFF"; CINZA_CLARO="F2F2F2"
HEADERS = {"Content-Type":"application/json","Accept":"application/json",
    "Origin":"https://comexstat.mdic.gov.br","Referer":"https://comexstat.mdic.gov.br/",
    "User-Agent":"Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"}

def _carregar_template():
    return _HTML_TEMPLATE

def _codigo_ncm(v):
    d = re.sub(r"\D","",str(v)); i = d.find("8701")
    return (d[i:i+8]).ljust(8,"0")[:8] if i>=0 else d[:8].ljust(8,"0")[:8]

def _segmento_por_ncm(cod):
    return PREFIXO_SEGMENTO.get(str(cod)[:6],("Outros",""))

def obter_ultimo_periodo(language="pt"):
    try:
        r = requests.get(f"{DATES_URL}?language={language}",headers=HEADERS,timeout=60,verify=False)
        r.raise_for_status(); data=r.json().get("data",r.json()); ano=mes=None
        if isinstance(data,dict):
            for k,v in data.items():
                kl=k.lower()
                if "ano" in kl or "year" in kl: ano=int(str(v)[:4])
                if "mes" in kl or "month" in kl: mes=int(v)
        if ano is None and isinstance(data,(str,list)):
            m=re.search(r"(20\d{2})[-/](\d{1,2})",str(data))
            if m: ano,mes=int(m.group(1)),int(m.group(2))
        if ano and mes:
            p=f"{ano:04d}-{mes:02d}"; print(f"[OK] Último período publicado: {p}"); return p
        raise ValueError("Formato de data não reconhecido.")
    except Exception as e:
        hoje=datetime.now(); ano,mes=hoje.year,hoje.month-2
        if mes<=0: mes+=12; ano-=1
        p=f"{ano:04d}-{mes:02d}"; print(f"[aviso] Data via API falhou ({e}). Estimando: {p}"); return p

def resolver_pasta_destino():
    if PASTA_DESTINO:
        p=Path(PASTA_DESTINO)
        try: p.mkdir(parents=True,exist_ok=True); return p
        except OSError: print("[aviso] Pasta configurada indisponível; usando Desktop.")
    home=Path.home(); cand=[home/"Desktop",home/"Área de Trabalho",home/"OneDrive"/"Desktop",home/"OneDrive"/"Área de Trabalho"]
    for od in home.glob("OneDrive*"): cand+=[od/"Desktop",od/"Área de Trabalho"]
    for c in cand:
        if c.exists(): return c
    fb=home/"Desktop"; fb.mkdir(parents=True,exist_ok=True); return fb

def consultar_importacao_tratores(period_from,period_to,ncms=None,language="pt",max_retries=5,timeout=180):
    if ncms is None: ncms=NCMS_TRATORES
    payload={"flow":"import","monthDetail":True,"period":{"from":period_from,"to":period_to},
        "filters":[{"filter":"ncm","values":[int(n) for n in ncms]}],"details":["ncm","country"],
        "metrics":["metricStatistic","metricFOB","metricCIF"]}
    url=f"{BASE_URL}?language={language}"
    for attempt in range(1,max_retries+1):
        try:
            resp=requests.post(url,headers=HEADERS,data=json.dumps(payload),timeout=timeout)
            resp.raise_for_status(); body=resp.json()
            if not body.get("success",False): raise RuntimeError(f"API erro: {body.get('message')}")
            df=pd.DataFrame(body.get("data",{}).get("list",[]))
            print(f"[OK] {len(df)} registros ({period_from} a {period_to})."); return _preparar(df)
        except requests.exceptions.SSLError:
            print("[aviso] SSL - tentando sem verificação...")
            resp=requests.post(url,headers=HEADERS,data=json.dumps(payload),timeout=timeout,verify=False)
            resp.raise_for_status(); return _preparar(pd.DataFrame(resp.json().get("data",{}).get("list",[])))
        except (requests.exceptions.RequestException,ValueError) as e:
            espera=2**attempt; print(f"[tentativa {attempt}/{max_retries}] {e}. Aguardando {espera}s..."); time.sleep(espera)
    raise RuntimeError("Falha ao obter dados.")

def _preparar(df):
    if df.empty: return df
    for col in ("metricStatistic","metricFOB","metricCIF","year","monthNumber"):
        if col in df.columns: df[col]=pd.to_numeric(df[col],errors="coerce")
    col_ano=next((c for c in ("year","coAno","ano") if c in df.columns),None)
    col_mes=next((c for c in ("monthNumber","coMes","mes","month") if c in df.columns),None)
    if col_ano: df["ano"]=pd.to_numeric(df[col_ano],errors="coerce").astype("Int64")
    if col_mes: df["mes"]=pd.to_numeric(df[col_mes],errors="coerce").astype("Int64")
    col_ncm=next((c for c in ("coNcm","CO_NCM","co_ncm","ncm","NCM","noNcm","noNcmpt") if c in df.columns),None)
    if col_ncm:
        cod=df[col_ncm].map(_codigo_ncm); seg=cod.map(_segmento_por_ncm)
        df["ncm"]=cod; df["segmento_hp"]=seg.map(lambda t:t[0]); df["descricao"]=seg.map(lambda t:t[1])
    col_pais=next((c for c in ("country","noPaispt","coPais","País","pais") if c in df.columns),None)
    if col_pais and col_pais!="pais": df=df.rename(columns={col_pais:"pais"})
    df=df.rename(columns={"metricStatistic":"quantidade_un","metricFOB":"valor_fob_usd","metricCIF":"valor_cif_usd"})
    if "valor_fob_usd" in df.columns and "quantidade_un" in df.columns:
        df["preco_medio_fob_un"]=(df["valor_fob_usd"]/df["quantidade_un"].replace(0,pd.NA)).round(0)
    cols=["ano","mes","pais","ncm","segmento_hp","descricao","quantidade_un","valor_fob_usd","valor_cif_usd","preco_medio_fob_un"]
    cols=[c for c in cols if c in df.columns]
    sc=[c for c in ("ano","mes","segmento_hp","valor_fob_usd") if c in df.columns]
    asc=[True,True,True,False][:len(sc)]
    return df[cols].sort_values(sc,ascending=asc).reset_index(drop=True)

def exportar_excel(df,caminho):
    if df.empty: print("[aviso] DF vazio."); return
    wb=Workbook(); ws=wb.active; ws.title="Detalhe"
    _escrever_aba(ws,df,"IMPORTAÇÃO DE TRATORES POR ANO/MÊS, PAÍS E SEGMENTO DE POTÊNCIA",
        ["Ano","Mês","País","NCM","Segmento (HP)","Descrição","Quantidade (un.)","Valor FOB (US$)","Valor CIF (US$)","Preço Médio FOB/un. (US$)"])
    resumo=(df.groupby("segmento_hp",as_index=False).agg(quantidade_un=("quantidade_un","sum"),valor_fob_usd=("valor_fob_usd","sum"),valor_cif_usd=("valor_cif_usd","sum")))
    resumo["preco_medio_fob_un"]=(resumo["valor_fob_usd"]/resumo["quantidade_un"].replace(0,pd.NA)).round(0)
    resumo["__o"]=resumo["segmento_hp"].map({s:i for i,s in enumerate(ORDEM_HP)}).fillna(99)
    resumo=resumo.sort_values("__o").drop(columns="__o").reset_index(drop=True)
    ws2=wb.create_sheet("Resumo por Segmento")
    _escrever_aba(ws2,resumo,"RESUMO POR SEGMENTO DE POTÊNCIA",["Segmento (HP)","Quantidade (un.)","Valor FOB (US$)","Valor CIF (US$)","Preço Médio FOB/un. (US$)"],total=True)
    matriz=pd.pivot_table(df,index="pais",columns="segmento_hp",values="valor_fob_usd",aggfunc="sum",fill_value=0)
    matriz=matriz.reindex(columns=[c for c in ORDEM_HP if c in matriz.columns]); matriz["Total"]=matriz.sum(axis=1)
    matriz=matriz.sort_values("Total",ascending=False).reset_index()
    ws3=wb.create_sheet("Matriz País x Segmento"); _escrever_matriz(ws3,matriz)
    wb.save(caminho); print(f"[OK] Excel salvo em: {caminho}")

def gerar_html_do_excel(caminho_excel,caminho_html,period_from="",period_to=""):
    dfx=pd.read_excel(caminho_excel,sheet_name="Detalhe",skiprows=1)
    ren={"Ano":"ano","Mês":"mes","País":"pais","NCM":"ncm","Segmento (HP)":"segmento_hp","Descrição":"descricao",
        "Quantidade (un.)":"quantidade_un","Valor FOB (US$)":"valor_fob_usd","Valor CIF (US$)":"valor_cif_usd","Preço Médio FOB/un. (US$)":"preco_medio_fob_un"}
    dfx=dfx.rename(columns=ren); dfx=dfx[[c for c in ren.values() if c in dfx.columns]].dropna(subset=["pais"])
    dfx["ncm"]=dfx["ncm"].astype(str).str.replace(r"\.0$","",regex=True)
    registros=dfx.where(pd.notna(dfx),None).to_dict(orient="records")
    meta={"period_from":period_from,"period_to":period_to,"gerado_em":datetime.now().strftime("%d/%m/%Y %H:%M"),"total_registros":len(registros)}
    html=(_carregar_template().replace("/*__META__*/null",json.dumps(meta,ensure_ascii=False)).replace("/*__DADOS__*/null",json.dumps(registros,ensure_ascii=False)))
    Path(caminho_html).write_text(html,encoding="utf-8"); print(f"[OK] Dashboard HTML gerado: {caminho_html}")

def _escrever_aba(ws,df,titulo,cabecalhos,total=False):
    borda=Border(*(Side(style="thin",color="D9D9D9"),)*4); n_col=len(cabecalhos)
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n_col)
    c=ws.cell(row=1,column=1,value=titulo); c.font=Font(name="Arial",size=13,bold=True,color=BRANCO)
    c.fill=PatternFill("solid",fgColor=AGCO_VERMELHO); c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=26
    for j,h in enumerate(cabecalhos,start=1):
        c=ws.cell(row=2,column=j,value=h); c.font=Font(name="Arial",size=10,bold=True,color=BRANCO)
        c.fill=PatternFill("solid",fgColor=AGCO_CINZA); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=borda
    ws.row_dimensions[2].height=32
    money={"valor_fob_usd","valor_cif_usd","preco_medio_fob_un"}; qty={"quantidade_un"}; inx={"ano","mes"}
    for i,(_,row) in enumerate(df.iterrows(),start=3):
        for j,col in enumerate(df.columns,start=1):
            c=ws.cell(row=i,column=j,value=row[col]); c.font=Font(name="Arial",size=10); c.border=borda
            if i%2==1: c.fill=PatternFill("solid",fgColor=CINZA_CLARO)
            if col in money: c.number_format='$#,##0'; c.alignment=Alignment(horizontal="right")
            elif col in qty: c.number_format='#,##0'; c.alignment=Alignment(horizontal="right")
            elif col in inx: c.number_format='0'; c.alignment=Alignment(horizontal="center")
            else: c.alignment=Alignment(horizontal="left")
    if total:
        r=len(df)+3; ws.cell(row=r,column=1,value="TOTAL").font=Font(name="Arial",size=10,bold=True)
        for j,col in enumerate(df.columns,start=1):
            letra=get_column_letter(j)
            if col in {"quantidade_un","valor_fob_usd","valor_cif_usd"}:
                c=ws.cell(row=r,column=j,value=f"=SUM({letra}3:{letra}{r-1})"); c.font=Font(name="Arial",size=10,bold=True)
                c.number_format='$#,##0' if col!="quantidade_un" else '#,##0'; c.alignment=Alignment(horizontal="right")
            elif col=="preco_medio_fob_un":
                fob=get_column_letter(list(df.columns).index("valor_fob_usd")+1); q=get_column_letter(list(df.columns).index("quantidade_un")+1)
                c=ws.cell(row=r,column=j,value=f"={fob}{r}/{q}{r}"); c.font=Font(name="Arial",size=10,bold=True); c.number_format='$#,##0'; c.alignment=Alignment(horizontal="right")
        for j in range(1,n_col+1): ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor="E8E8E8")
    larg={"ano":8,"mes":8,"pais":22,"ncm":12,"segmento_hp":20,"descricao":24,"quantidade_un":16,"valor_fob_usd":18,"valor_cif_usd":18,"preco_medio_fob_un":22}
    for j,col in enumerate(df.columns,start=1): ws.column_dimensions[get_column_letter(j)].width=larg.get(col,18)
    ws.freeze_panes="A3"; nota=ws.max_row+2
    ws.cell(row=nota,column=1,value="Fonte: Comex Stat / MDIC. Segmentação de potência: AGCO — Global Reporting & Analytics.").font=Font(name="Arial",size=8,italic=True,color=AGCO_CINZA)

def _escrever_matriz(ws,matriz):
    borda=Border(*(Side(style="thin",color="D9D9D9"),)*4); n_col=matriz.shape[1]
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=n_col)
    c=ws.cell(row=1,column=1,value="MATRIZ PAÍS x SEGMENTO — VALOR FOB (US$)"); c.font=Font(name="Arial",size=13,bold=True,color=BRANCO)
    c.fill=PatternFill("solid",fgColor=AGCO_VERMELHO); c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=26
    headers=["País"]+list(matriz.columns[1:])
    for j,h in enumerate(headers,start=1):
        c=ws.cell(row=2,column=j,value=h); c.font=Font(name="Arial",size=10,bold=True,color=BRANCO)
        c.fill=PatternFill("solid",fgColor=AGCO_CINZA); c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=borda
    ws.row_dimensions[2].height=32
    for i,(_,row) in enumerate(matriz.iterrows(),start=3):
        for j,col in enumerate(matriz.columns,start=1):
            c=ws.cell(row=i,column=j,value=row[col]); c.font=Font(name="Arial",size=10,bold=(col=="Total")); c.border=borda
            if i%2==1: c.fill=PatternFill("solid",fgColor=CINZA_CLARO)
            if j==1: c.alignment=Alignment(horizontal="left")
            else: c.number_format='$#,##0'; c.alignment=Alignment(horizontal="right")
    r=len(matriz)+3; ws.cell(row=r,column=1,value="TOTAL").font=Font(name="Arial",size=10,bold=True)
    for j in range(2,n_col+1):
        letra=get_column_letter(j); c=ws.cell(row=r,column=j,value=f"=SUM({letra}3:{letra}{r-1})")
        c.font=Font(name="Arial",size=10,bold=True); c.number_format='$#,##0'; c.alignment=Alignment(horizontal="right")
    for j in range(1,n_col+1): ws.cell(row=r,column=j).fill=PatternFill("solid",fgColor="E8E8E8")
    ws.column_dimensions["A"].width=22
    for j in range(2,n_col+1): ws.column_dimensions[get_column_letter(j)].width=20
    ws.freeze_panes="B3"

_HTML_TEMPLATE = r"""<!DOCTYPE html>
<html lang="pt-BR"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Importação de Tratores — Comex Stat</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
<style>
:root{--red:#C8102E;--red-dark:#9E0C24;--red-soft:#F6D5DB;--charcoal:#22252A;--gray:#5B616B;--gray-light:#8A909A;--bg:#F0F2F5;--card:#FFFFFF;--border:#E6E9ED;--green:#1E9E5A;--shadow:0 1px 3px rgba(20,24,31,.06),0 8px 24px rgba(20,24,31,.05);--shadow-lg:0 4px 12px rgba(20,24,31,.08),0 16px 40px rgba(20,24,31,.08)}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Inter',-apple-system,'Segoe UI',Roboto,Arial,sans-serif;background:var(--bg);color:var(--charcoal);font-size:14px;-webkit-font-smoothing:antialiased}
.topbar{background:var(--card);border-bottom:1px solid var(--border);padding:16px 32px;display:flex;align-items:center;justify-content:space-between;position:sticky;top:0;z-index:50;box-shadow:0 1px 2px rgba(20,24,31,.04)}
.title-wrap{display:flex;align-items:center;gap:14px}.title-accent{width:5px;height:38px;background:linear-gradient(180deg,var(--red),var(--red-dark));border-radius:3px}
.title-wrap h1{font-size:18px;font-weight:800;letter-spacing:-.3px}.title-wrap p{font-size:12px;color:var(--gray-light);font-weight:500;margin-top:1px}
.top-right{display:flex;align-items:center;gap:18px}.meta{text-align:right;font-size:11.5px;color:var(--gray-light);line-height:1.55}.meta b{color:var(--gray)}
.lang{display:flex;background:var(--bg);border:1px solid var(--border);border-radius:9px;padding:3px;gap:2px}
.lang button{border:none;background:transparent;padding:6px 12px;border-radius:6px;font-size:12px;font-weight:700;color:var(--gray-light);cursor:pointer;font-family:inherit;transition:.15s;letter-spacing:.3px}
.lang button:hover{color:var(--charcoal)}.lang button.on{background:var(--red);color:#fff;box-shadow:0 2px 6px rgba(200,16,46,.3)}
.wrap{max-width:1440px;margin:0 auto;padding:26px 32px 40px}
.sec{font-size:12px;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:var(--gray-light);margin:26px 0 13px;display:flex;align-items:center;gap:8px}
.sec::before{content:"";width:14px;height:2px;background:var(--red);border-radius:2px}
.kpis{display:grid;grid-template-columns:repeat(4,1fr);gap:18px}
.kpi{background:var(--card);border-radius:14px;padding:20px 22px;box-shadow:var(--shadow);position:relative;overflow:hidden;transition:.2s}
.kpi:hover{transform:translateY(-2px);box-shadow:var(--shadow-lg)}.kpi::before{content:"";position:absolute;left:0;top:0;bottom:0;width:4px;background:linear-gradient(180deg,var(--red),var(--red-dark))}
.kpi .label{font-size:11px;text-transform:uppercase;letter-spacing:.6px;color:var(--gray-light);font-weight:600}.kpi .value{font-size:28px;font-weight:800;margin-top:8px;letter-spacing:-.5px}.kpi .sub{font-size:11px;color:var(--gray-light);margin-top:3px;font-weight:500}
.filters{background:var(--card);border-radius:14px;padding:20px 22px;box-shadow:var(--shadow);display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:16px;align-items:end}
.fld{display:flex;flex-direction:column;gap:6px}.fld label{font-size:11px;font-weight:700;color:var(--gray);text-transform:uppercase;letter-spacing:.4px}
.fld select,.fld input{padding:9px 11px;border:1px solid var(--border);border-radius:9px;font-size:13px;font-family:inherit;background:#fff;color:var(--charcoal);transition:.15s}
.fld select:focus,.fld input:focus{outline:none;border-color:var(--red);box-shadow:0 0 0 3px var(--red-soft)}.fld select[multiple]{height:96px}
.btn{padding:10px 18px;border:none;border-radius:9px;font-size:13px;font-weight:700;cursor:pointer;font-family:inherit;transition:.15s;letter-spacing:.2px}
.btn-primary{background:var(--red);color:#fff;box-shadow:0 2px 8px rgba(200,16,46,.25)}.btn-primary:hover{background:var(--red-dark);transform:translateY(-1px)}
.btn-ghost{background:#fff;color:var(--gray);border:1px solid var(--border)}.btn-ghost:hover{background:var(--bg);border-color:var(--gray-light)}
.actions{display:flex;gap:11px;flex-wrap:wrap;margin-top:16px}
.card{background:var(--card);border-radius:14px;padding:20px 22px;box-shadow:var(--shadow)}
.card h3{font-size:13px;font-weight:700;margin-bottom:14px;letter-spacing:-.2px;display:flex;justify-content:space-between;align-items:center;gap:10px}
.card h3 .mini{font-size:11px;font-weight:600;color:var(--gray-light)}
.chart-box{position:relative;height:300px}.chart-box.tall{height:340px}
.seg-split{display:grid;grid-template-columns:1fr 1fr;gap:18px;align-items:center}
.filter-bar{display:flex;flex-wrap:wrap;gap:8px;margin:-4px 0 16px}
.filter-bar .fchip{display:inline-flex;align-items:center;gap:7px;background:var(--bg);border:1px solid var(--border);border-radius:8px;padding:6px 12px;font-size:12px;font-weight:600;color:var(--charcoal)}
.filter-bar .fchip b{color:var(--red);font-weight:700}
.filter-bar .fchip .ic{font-size:13px;opacity:.75}
.mini-wrap{max-height:640px;overflow:auto;border:1px solid var(--border);border-radius:10px}
.mini-table td.stick,.mini-table th.stick{position:sticky;left:0;z-index:3;background:#fff;box-shadow:2px 0 0 var(--border)}
.mini-table thead th.stick{z-index:5;background:var(--charcoal)}
.mini-table.matrix thead tr.grp th.stick{background:#1A1D22}
.mini-table tbody tr:hover td.stick{background:#FFF5F6}
.mini-table tbody tr.sel td.stick{background:var(--red-soft)!important}
.mini-table{width:100%;border-collapse:collapse;font-size:12.5px}
.mini-table thead th{background:var(--charcoal);color:#fff;padding:9px 11px;text-align:center;font-weight:600;font-size:11.5px;cursor:pointer;white-space:nowrap;position:sticky;top:0}
.mini-table.matrix thead tr.grp th{background:#1A1D22;border-left:2px solid #3A3E45;font-size:12px;letter-spacing:.3px}
.mini-table.matrix thead tr.grp th.tot{background:var(--red-dark)}.mini-table.matrix thead tr.grp th.var{background:#111}
.mini-table.matrix thead tr.sub th{background:var(--charcoal);font-size:10.5px;font-weight:600;top:32px}.mini-table.matrix thead tr.sub th.b{border-left:2px solid #3A3E45}
.mini-table tbody td{padding:8px 11px;border-bottom:1px solid var(--border);white-space:nowrap;text-align:right;font-variant-numeric:tabular-nums}
.mini-table tbody td.nm{text-align:left;font-weight:600;cursor:pointer}.mini-table tbody td.b{border-left:2px solid var(--border)}.mini-table tbody td.tot{background:#FBECEF;font-weight:700}
.mini-table tbody tr:hover td{background:#FFF5F6}.mini-table tbody tr.sel td{background:var(--red-soft)!important}
.var-up{color:var(--green);font-weight:700}.var-dn{color:var(--red);font-weight:700}.var-0{color:var(--gray-light)}
.seg-stats{display:flex;flex-direction:column;gap:9px;justify-content:center}
.seg-stats .row{display:flex;align-items:center;justify-content:space-between;gap:14px;font-size:12.5px;padding:7px 12px;border-radius:9px;background:var(--bg)}
.seg-stats .row .nm{display:flex;align-items:center;gap:8px;font-weight:600}.seg-stats .row .dot{width:11px;height:11px;border-radius:3px;flex:none}
.seg-stats .row .vl{font-variant-numeric:tabular-nums;font-weight:700}.seg-stats .row .pc{font-size:11px;color:var(--gray-light);font-weight:600;min-width:44px;text-align:right}
.chip{display:inline-flex;align-items:center;gap:6px;background:var(--red-soft);color:var(--red-dark);padding:4px 12px;border-radius:20px;font-size:12px;font-weight:700;cursor:pointer}.chip span{font-size:14px;line-height:1}
.hint{font-size:11.5px;color:var(--gray-light);margin-top:9px;font-weight:500}
.table-wrap{background:var(--card);border-radius:14px;box-shadow:var(--shadow);overflow:hidden}.table-scroll{overflow-x:auto;max-height:540px;overflow-y:auto}
table.main{width:100%;border-collapse:collapse;font-size:13px}
table.main thead th{background:var(--charcoal);color:#fff;padding:12px 13px;text-align:left;position:sticky;top:0;cursor:pointer;white-space:nowrap;font-weight:600;user-select:none;transition:.15s;font-size:12px;letter-spacing:.2px}
table.main thead th:hover{background:#33373E}table.main thead th.num{text-align:right}
table.main tbody td{padding:10px 13px;border-bottom:1px solid var(--border);white-space:nowrap}table.main tbody td.num{text-align:right;font-variant-numeric:tabular-nums}table.main tbody td.ctr{text-align:center}
table.main tbody tr:nth-child(even){background:#FAFBFC}table.main tbody tr:hover{background:#FFF5F6}
table.main tfoot td{padding:12px 13px;font-weight:800;background:#F5F6F8;border-top:2px solid var(--red)}table.main tfoot td.num{text-align:right;font-variant-numeric:tabular-nums}
.seg-tag{display:inline-block;padding:3px 10px;border-radius:20px;font-size:11px;font-weight:700;color:#fff;letter-spacing:.2px}
footer{text-align:center;color:var(--gray-light);font-size:12px;padding:28px;line-height:1.7}footer strong{color:var(--red)}
@media(max-width:900px){.kpis,.filters,.seg-split{grid-template-columns:1fr}.topbar{flex-direction:column;gap:12px;align-items:flex-start}}
</style></head><body>
<div class="topbar"><div class="title-wrap"><div class="title-accent"></div>
<div><h1 data-i18n="title">Importação Brasileira de Tratores</h1><p data-i18n="subtitle">Análise por país, ano/mês e segmento de potência (HP) · Comex Stat</p></div></div>
<div class="top-right"><div class="meta" id="meta"></div><div class="lang">
<button data-lang="pt" class="on" onclick="setLang('pt')">PT</button><button data-lang="en" onclick="setLang('en')">EN</button><button data-lang="es" onclick="setLang('es')">ES</button></div></div></div>
<div class="wrap">
<div class="sec" data-i18n="sec_kpi">Indicadores</div><div class="kpis" id="kpis"></div>
<div class="sec" data-i18n="sec_filters">Filtros</div>
<div class="filters">
<div class="fld"><label data-i18n="f_year_from">Ano — de</label><select id="fAnoDe" onchange="aplicar()"></select></div>
<div class="fld"><label data-i18n="f_year_to">Ano — até</label><select id="fAnoAte" onchange="aplicar()"></select></div>
<div class="fld"><label data-i18n="f_month">Mês</label><select id="fMes" multiple onchange="aplicar()"></select></div>
<div class="fld"><label data-i18n="f_country">País (Ctrl p/ vários)</label><select id="fPais" multiple onchange="aplicar()"></select></div>
<div class="fld"><label data-i18n="f_segment">Segmento (HP)</label><select id="fSeg" multiple onchange="aplicar()"></select></div>
<div class="fld"><label data-i18n="f_var">Variável</label><select id="fVar" onchange="aplicar()">
<option value="all" data-i18n="v_all">Todas</option><option value="quantidade_un" data-i18n="v_qty">Quantidade</option>
<option value="valor_fob_usd" data-i18n="v_fob">Valor FOB</option><option value="valor_cif_usd" data-i18n="v_cif">Valor CIF</option>
<option value="preco_medio_fob_un" data-i18n="v_avg">Preço Médio</option></select></div>
<div class="fld"><label data-i18n="f_search">Buscar (país/NCM)</label><input type="text" id="fBusca" placeholder="China, 87019590…" oninput="aplicar()"></div>
</div>
<div class="actions"><button class="btn btn-primary" data-i18n="b_apply" onclick="aplicar()">Aplicar filtros</button>
<button class="btn btn-ghost" data-i18n="b_clear" onclick="limpar()">Limpar</button>
<button class="btn btn-ghost" data-i18n="b_csv" onclick="exportarCSV()">⬇ Exportar CSV</button></div>
<div class="sec" data-i18n="sec_country">Volume por país e segmento</div>
<div class="card"><h3><span data-i18n="c_ctable">Volume por país e ano</span> <span class="mini" data-i18n="c_ctable_hint">· clique num país para filtrar a rosca</span></h3>
<div class="mini-wrap"><table class="mini-table matrix" id="tbPais"><thead id="tbPaisHead"></thead><tbody id="tbPaisBody"></tbody></table></div><div class="hint" id="varHint"></div></div>
<div class="card" style="margin-top:18px"><h3><span id="segTitle" data-i18n="c_cseg">Segmentos — Todos os países</span> <span id="segChip"></span></h3>
<div id="segFilters" class="filter-bar"></div>
<div class="seg-split"><div class="chart-box tall"><canvas id="chPaisSeg"></canvas></div><div id="segStats" class="seg-stats"></div></div><div class="hint" id="segAvg"></div></div>
<div class="sec"><span data-i18n="sec_table">Tabela de dados</span> <span id="cnt" style="font-weight:400;text-transform:none;letter-spacing:0;color:var(--gray-light)"></span></div>
<div class="table-wrap"><div class="table-scroll"><table class="main"><thead><tr id="thead"></tr></thead><tbody id="tbody"></tbody><tfoot id="tfoot"></tfoot></table></div></div>
</div>
<footer><span data-i18n="foot1">Metodologia de segmentação por potência desenvolvida pela área de</span> <strong>Global Reporting &amp; Analytics</strong> <span data-i18n="foot2">através de</span> <strong>Thiago Montoro</strong>.<br><span data-i18n="foot3">Fonte dos dados</span>: <strong>Comex Stat / MDIC</strong></footer>
<script>
const COMEX_META=/*__META__*/null;const COMEX_DATA=/*__DADOS__*/null;
const CORES_SEG={"Até 24 HP":"#F4A6B0","Acima de 24 a 50 HP":"#E8607A","Acima de 50 a 100 HP":"#C8102E","Acima de 101 a 175 HP":"#8E1B3B","Acima de 175 HP":"#53565A","Outros":"#B0B4B8"};
const ORDEM_SEG=["Até 24 HP","Acima de 24 a 50 HP","Acima de 50 a 100 HP","Acima de 101 a 175 HP","Acima de 175 HP","Outros"];
const MES_I18N={pt:["Jan","Fev","Mar","Abr","Mai","Jun","Jul","Ago","Set","Out","Nov","Dez"],en:["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"],es:["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]};
const SEG_I18N={pt:{"Até 24 HP":"Até 24 HP","Acima de 24 a 50 HP":"24 a 50 HP","Acima de 50 a 100 HP":"50 a 100 HP","Acima de 101 a 175 HP":"101 a 175 HP","Acima de 175 HP":"Acima de 175 HP","Outros":"Outros"},en:{"Até 24 HP":"Up to 24 HP","Acima de 24 a 50 HP":"24 to 50 HP","Acima de 50 a 100 HP":"50 to 100 HP","Acima de 101 a 175 HP":"101 to 175 HP","Acima de 175 HP":"Above 175 HP","Outros":"Others"},es:{"Até 24 HP":"Hasta 24 HP","Acima de 24 a 50 HP":"24 a 50 HP","Acima de 50 a 100 HP":"50 a 100 HP","Acima de 101 a 175 HP":"101 a 175 HP","Acima de 175 HP":"Más de 175 HP","Outros":"Otros"}};
const DESC_I18N={pt:{"Micro/compactos":"Micro/compactos","Pequenos":"Pequenos","Médios":"Médios","Grandes":"Grandes","Alta Potência":"Alta Potência","Outros":"Outros"},en:{"Micro/compactos":"Micro/compact","Pequenos":"Small","Médios":"Medium","Grandes":"Large","Alta Potência":"High power","Outros":"Others"},es:{"Micro/compactos":"Micro/compactos","Pequenos":"Pequeños","Médios":"Medianos","Grandes":"Grandes","Alta Potência":"Alta potencia","Outros":"Otros"}};
const PAIS_I18N={en:{"China":"China","Índia":"India","India":"India","Estados Unidos":"United States","Itália":"Italy","Finlândia":"Finland","Alemanha":"Germany","Turquia":"Turkey","Suécia":"Sweden","França":"France","Coreia do Sul":"South Korea","Espanha":"Spain","Brasil":"Brazil","Canadá":"Canada","México":"Mexico","Argentina":"Argentina","Outros":"Others"},es:{"China":"China","Índia":"India","India":"India","Estados Unidos":"Estados Unidos","Itália":"Italia","Finlândia":"Finlandia","Alemanha":"Alemania","Turquia":"Turquía","Suécia":"Suecia","França":"Francia","Coreia do Sul":"Corea del Sur","Espanha":"España","Brasil":"Brasil","Canadá":"Canadá","México":"México","Argentina":"Argentina","Outros":"Otros"}};
const I18N={
pt:{title:"Importação Brasileira de Tratores",subtitle:"Análise por país, ano/mês e segmento de potência (HP) · Comex Stat",sec_kpi:"Indicadores",sec_filters:"Filtros",sec_country:"Volume por país e segmento",sec_table:"Tabela de dados",f_year_from:"Ano — de",f_year_to:"Ano — até",f_month:"Mês",f_country:"País (Ctrl p/ vários)",f_segment:"Segmento (HP)",f_search:"Buscar (país/NCM)",f_var:"Variável",v_all:"Todas",v_qty:"Quantidade",v_fob:"Valor FOB",v_cif:"Valor CIF",v_avg:"Preço Médio",b_apply:"Aplicar filtros",b_clear:"Limpar",b_csv:"⬇ Exportar CSV",c_ctable:"Volume por país e ano",c_ctable_hint:"· clique num país para filtrar a rosca",c_cseg:"Segmentos — Todos os países",c_seg_of:"Segmentos —",all_countries:"Todos os países",fb_country:"País",fb_year:"Ano",fb_month:"Acumulado",fb_var:"Variável",m_all:"Jan → Dez",k_fob:"Valor FOB total",k_cif:"Valor CIF total",k_units:"Unidades importadas",k_avg:"Preço médio",k_fob_s:"US$ (Free On Board)",k_cif_s:"US$ (Custo, Seguro e Frete)",k_units_s:"tratores",k_avg_s:"FOB por unidade",th_year:"Ano",th_month:"Mês",th_country:"País",th_ncm:"NCM",th_seg:"Segmento (HP)",th_desc:"Descrição",th_qty:"Qtd (un.)",th_fob:"FOB (US$)",th_cif:"CIF (US$)",th_avg:"Preço Médio (US$)",th_var:"Var% YoY",total:"TOTAL",lines:"linhas",rowsmore:"exibindo {a} de {b} linhas. Use os filtros para refinar.",avg_txt:"Preço médio {c}: {v} / unidade",tot_txt:"{lbl} {c}: {v}",var_txt:"Var% interanual = {met} {y1} vs {y0}, no mesmo acumulado (Jan → {m}).",foot1:"Metodologia de segmentação por potência desenvolvida pela área de",foot2:"através de",foot3:"Fonte dos dados",period:"Período",updated:"Atualizado",records:"registros"},
en:{title:"Brazilian Tractor Imports",subtitle:"Analysis by country, year/month and power segment (HP) · Comex Stat",sec_kpi:"Key indicators",sec_filters:"Filters",sec_country:"Volume by country and segment",sec_table:"Data table",f_year_from:"Year — from",f_year_to:"Year — to",f_month:"Month",f_country:"Country (Ctrl for many)",f_segment:"Segment (HP)",f_search:"Search (country/HS)",f_var:"Variable",v_all:"All",v_qty:"Quantity",v_fob:"FOB value",v_cif:"CIF value",v_avg:"Avg. price",b_apply:"Apply filters",b_clear:"Clear",b_csv:"⬇ Export CSV",c_ctable:"Volume by country and year",c_ctable_hint:"· click a country to filter the doughnut",c_cseg:"Segments — All countries",c_seg_of:"Segments —",all_countries:"All countries",fb_country:"Country",fb_year:"Year",fb_month:"Cumulative",fb_var:"Variable",m_all:"Jan → Dec",k_fob:"Total FOB value",k_cif:"Total CIF value",k_units:"Units imported",k_avg:"Average price",k_fob_s:"US$ (Free On Board)",k_cif_s:"US$ (Cost, Insurance, Freight)",k_units_s:"tractors",k_avg_s:"FOB per unit",th_year:"Year",th_month:"Month",th_country:"Country",th_ncm:"HS code",th_seg:"Segment (HP)",th_desc:"Description",th_qty:"Qty (units)",th_fob:"FOB (US$)",th_cif:"CIF (US$)",th_avg:"Avg. price (US$)",th_var:"YoY %",total:"TOTAL",lines:"rows",rowsmore:"showing {a} of {b} rows. Use filters to refine.",avg_txt:"Average price {c}: {v} / unit",tot_txt:"{lbl} {c}: {v}",var_txt:"YoY % = {met} {y1} vs {y0}, same cumulative window (Jan → {m}).",foot1:"Power segmentation methodology developed by the",foot2:"through",foot3:"Data source",period:"Period",updated:"Updated",records:"records"},
es:{title:"Importación Brasileña de Tractores",subtitle:"Análisis por país, año/mes y segmento de potencia (HP) · Comex Stat",sec_kpi:"Indicadores",sec_filters:"Filtros",sec_country:"Volumen por país y segmento",sec_table:"Tabla de datos",f_year_from:"Año — desde",f_year_to:"Año — hasta",f_month:"Mes",f_country:"País (Ctrl para varios)",f_segment:"Segmento (HP)",f_search:"Buscar (país/NCM)",f_var:"Variable",v_all:"Todas",v_qty:"Cantidad",v_fob:"Valor FOB",v_cif:"Valor CIF",v_avg:"Precio Medio",b_apply:"Aplicar filtros",b_clear:"Limpiar",b_csv:"⬇ Exportar CSV",c_ctable:"Volumen por país y año",c_ctable_hint:"· clic en un país para filtrar la rosca",c_cseg:"Segmentos — Todos los países",c_seg_of:"Segmentos —",all_countries:"Todos los países",fb_country:"País",fb_year:"Año",fb_month:"Acumulado",fb_var:"Variable",m_all:"Ene → Dic",k_fob:"Valor FOB total",k_cif:"Valor CIF total",k_units:"Unidades importadas",k_avg:"Precio medio",k_fob_s:"US$ (Free On Board)",k_cif_s:"US$ (Costo, Seguro y Flete)",k_units_s:"tractores",k_avg_s:"FOB por unidad",th_year:"Año",th_month:"Mes",th_country:"País",th_ncm:"NCM",th_seg:"Segmento (HP)",th_desc:"Descripción",th_qty:"Cant. (un.)",th_fob:"FOB (US$)",th_cif:"CIF (US$)",th_avg:"Precio Medio (US$)",th_var:"Var% interanual",total:"TOTAL",lines:"filas",rowsmore:"mostrando {a} de {b} filas. Use los filtros para refinar.",avg_txt:"Precio medio {c}: {v} / unidad",tot_txt:"{lbl} {c}: {v}",var_txt:"Var% interanual = {met} {y1} vs {y0}, misma ventana acumulada (Ene → {m}).",foot1:"Metodología de segmentación por potencia desarrollada por el área de",foot2:"a través de",foot3:"Fuente de los datos",period:"Período",updated:"Actualizado",records:"registros"}};
let LANG="pt",DADOS=[],FILTRADO=[],sortState={col:"ano",dir:1},paisSort={col:"qty",dir:-1},paisSortAno=null,paisSel=null;const charts={};
function t(k){return (I18N[LANG]&&I18N[LANG][k])||I18N.pt[k]||k;}
function segLabel(s){return (SEG_I18N[LANG]&&SEG_I18N[LANG][s])||s;}
function descLabel(d){return (DESC_I18N[LANG]&&DESC_I18N[LANG][d])||d;}
function paisLabel(p){if(LANG==="pt")return p;return (PAIS_I18N[LANG]&&PAIS_I18N[LANG][p])||p;}
function mesNome(m){return (MES_I18N[LANG]||MES_I18N.pt)[m-1]||m;}
const COLS=[{k:"ano",i:"th_year",ctr:true},{k:"mes",i:"th_month",ctr:true},{k:"pais",i:"th_country",tr:"pais"},{k:"ncm",i:"th_ncm"},{k:"segmento_hp",i:"th_seg",tag:true},{k:"descricao",i:"th_desc",tr:"desc"},{k:"quantidade_un",i:"th_qty",num:true,fmt:"int"},{k:"valor_fob_usd",i:"th_fob",num:true,fmt:"usd"},{k:"valor_cif_usd",i:"th_cif",num:true,fmt:"usd"},{k:"preco_medio_fob_un",i:"th_avg",num:true,fmt:"usd"}];
const MET_DEF=[{k:"quantidade_un",i:"th_qty",fmt:"int",kpi:"k_units"},{k:"valor_fob_usd",i:"th_fob",fmt:"usd",kpi:"k_fob"},{k:"valor_cif_usd",i:"th_cif",fmt:"usd",kpi:"k_cif"},{k:"preco_medio_fob_un",i:"th_avg",fmt:"usd",kpi:"k_avg"}];
function varSel(){return (document.getElementById("fVar")||{}).value||"all";}
function metsAtivas(){const v=varSel();return v==="all"?MET_DEF.slice():MET_DEF.filter(m=>m.k===v);}
// Métrica que a ROSCA usa: se "Todas" ou "Preço Médio" -> FOB; senão a própria escolhida
function metRosca(){const v=varSel();const m=MET_DEF.find(x=>x.k===v);return (v==="all"||v==="preco_medio_fob_un"||!m)?MET_DEF[1]:m;}
function fmtMet(m,val){return m.fmt==="usd"?fUsd(val):fInt(val);}
function fmtMetC(m,val){return m.fmt==="usd"?fUsdC(val):fInt(val);}
function colsVisiveis(){const mets=metsAtivas().map(m=>m.k);return COLS.filter(c=>!c.num||mets.includes(c.k));}
const loc=()=>LANG==="pt"?"pt-BR":LANG==="es"?"es-ES":"en-US";
const fInt=n=>(n==null||isNaN(n))?"—":Math.round(n).toLocaleString(loc());
const fUsd=n=>(n==null||isNaN(n))?"—":"$"+Math.round(n).toLocaleString(loc());
const fUsdC=n=>(n==null||isNaN(n))?"—":n>=1e9?"$"+(n/1e9).toFixed(2)+" bi":n>=1e6?"$"+(n/1e6).toFixed(1)+" mi":"$"+fInt(n);
function normalizar(a){return a.map(r=>{const ano=+r.ano,mes=+r.mes;return{ano,mes,periodo:ano+"-"+String(mes).padStart(2,"0"),pais:r.pais||"",ncm:String(r.ncm||""),segmento_hp:r.segmento_hp||"Outros",descricao:r.descricao||"",quantidade_un:+r.quantidade_un||0,valor_fob_usd:+r.valor_fob_usd||0,valor_cif_usd:+r.valor_cif_usd||0,preco_medio_fob_un:+r.preco_medio_fob_un||0};});}
window.addEventListener("DOMContentLoaded",()=>{DADOS=normalizar(COMEX_DATA||[]);popularFiltros();setLang("pt");aplicar();});
function setLang(l){LANG=l;document.documentElement.lang=l==="pt"?"pt-BR":l;
document.querySelectorAll(".lang button").forEach(b=>b.classList.toggle("on",b.dataset.lang===l));
document.querySelectorAll("[data-i18n]").forEach(el=>{el.textContent=t(el.dataset.i18n);});
[...fPais.options].forEach(o=>o.text=paisLabel(o.value));[...fSeg.options].forEach(o=>o.text=segLabel(o.value));[...fMes.options].forEach(o=>o.text=o.value+" - "+mesNome(+o.value));
renderMeta();render();}
function renderMeta(){if(!COMEX_META)return;document.getElementById("meta").innerHTML=t("period")+": <b>"+COMEX_META.period_from+" → "+COMEX_META.period_to+"</b><br>"+t("updated")+": "+COMEX_META.gerado_em+" · "+COMEX_META.total_registros+" "+t("records");}
function popularFiltros(){const anos=[...new Set(DADOS.map(d=>d.ano))].filter(v=>!isNaN(v)).sort((a,b)=>a-b);const meses=[...new Set(DADOS.map(d=>d.mes))].filter(v=>!isNaN(v)).sort((a,b)=>a-b);const paises=[...new Set(DADOS.map(d=>d.pais))].filter(Boolean).sort();const segs=ORDEM_SEG.filter(s=>DADOS.some(d=>d.segmento_hp===s));
fAnoDe.innerHTML="";fAnoAte.innerHTML="";anos.forEach(a=>{fAnoDe.add(new Option(a,a));fAnoAte.add(new Option(a,a));});fAnoDe.value=anos[0];fAnoAte.value=anos[anos.length-1];
fMes.innerHTML="";meses.forEach(m=>{const o=new Option(m+" - "+mesNome(m),m);o.selected=true;fMes.add(o);});
fPais.innerHTML="";paises.forEach(p=>fPais.add(new Option(paisLabel(p),p)));fSeg.innerHTML="";segs.forEach(s=>fSeg.add(new Option(segLabel(s),s)));}
function mSel(id){return[...document.getElementById(id).selectedOptions].map(o=>o.value);}
function aplicar(){const ad=+fAnoDe.value,aa=+fAnoAte.value,ms=mSel("fMes").map(Number),ps=mSel("fPais"),sg=mSel("fSeg"),bu=fBusca.value.trim().toLowerCase();
FILTRADO=DADOS.filter(d=>{if(d.ano<ad||d.ano>aa)return false;if(ms.length&&!ms.includes(d.mes))return false;if(ps.length&&!ps.includes(d.pais))return false;if(sg.length&&!sg.includes(d.segmento_hp))return false;if(bu&&!((d.pais+" "+paisLabel(d.pais)+" "+d.ncm+" "+d.segmento_hp).toLowerCase().includes(bu)))return false;return true;});paisSel=null;render();}
function limpar(){[...fMes.options].forEach(o=>o.selected=true);fPais.selectedIndex=-1;fSeg.selectedIndex=-1;fBusca.value="";const fv=document.getElementById("fVar");if(fv)fv.value="all";const anos=[...new Set(DADOS.map(d=>d.ano))].filter(v=>!isNaN(v)).sort((a,b)=>a-b);fAnoDe.value=anos[0];fAnoAte.value=anos[anos.length-1];aplicar();}
function render(){renderKPIs();renderTabela();renderPais();}
function soma(a,k){return a.reduce((s,d)=>s+(d[k]||0),0);}
function agr(a,k,v){const m={};a.forEach(d=>{m[d[k]]=(m[d[k]]||0)+(d[v]||0);});return m;}
function renderKPIs(){const fob=soma(FILTRADO,"valor_fob_usd"),cif=soma(FILTRADO,"valor_cif_usd"),qtd=soma(FILTRADO,"quantidade_un"),pm=qtd?fob/qtd:0;const k=[[t("k_fob"),fUsdC(fob),t("k_fob_s")],[t("k_cif"),fUsdC(cif),t("k_cif_s")],[t("k_units"),fInt(qtd),t("k_units_s")],[t("k_avg"),fUsd(pm),t("k_avg_s")]];kpis.innerHTML=k.map(x=>'<div class="kpi"><div class="label">'+x[0]+'</div><div class="value">'+x[1]+'</div><div class="sub">'+x[2]+'</div></div>').join("");}
function anosDisp(){return [...new Set(FILTRADO.map(d=>d.ano))].filter(v=>!isNaN(v)).sort((a,b)=>a-b);}
function celVal(c,m){if(m.k==="quantidade_un")return c.q;if(m.k==="valor_fob_usd")return c.f;if(m.k==="valor_cif_usd")return c.c;if(m.k==="preco_medio_fob_un")return c.q?c.f/c.q:0;return 0;}
function renderPais(){const anos=anosDisp();const mets=metsAtivas();const nm=mets.length;
const M={};FILTRADO.forEach(d=>{if(!M[d.pais])M[d.pais]={tot:{q:0,f:0,c:0}};if(!M[d.pais][d.ano])M[d.pais][d.ano]={q:0,f:0,c:0};const cel=M[d.pais][d.ano];cel.q+=d.quantidade_un;cel.f+=d.valor_fob_usd;cel.c+=d.valor_cif_usd;M[d.pais].tot.q+=d.quantidade_un;M[d.pais].tot.f+=d.valor_fob_usd;M[d.pais].tot.c+=d.valor_cif_usd;});
const ultimo=anos[anos.length-1],anterior=anos[anos.length-2];const maxMes=FILTRADO.filter(d=>d.ano===ultimo).length?Math.max(...FILTRADO.filter(d=>d.ano===ultimo).map(d=>d.mes)):12;
const vv=varSel();const kVar=(vv==="quantidade_un"||vv==="valor_cif_usd")?vv:"valor_fob_usd";const metVar=MET_DEF.find(m=>m.k===kVar);
const varYtd=(p,y)=>FILTRADO.filter(d=>d.pais===p&&d.ano===y&&d.mes<=maxMes).reduce((s,d)=>s+(d[kVar]||0),0);
let arr=Object.keys(M).map(p=>({pais:p,cel:M[p],v1:varYtd(p,ultimo),v0:anterior?varYtd(p,anterior):0}));
// Ano usado para ordenar por quantidade (padrão = último ano disponível, ex.: 2026)
const anoOrd=(paisSortAno&&anos.includes(paisSortAno))?paisSortAno:ultimo;
const qtdAno=r=>((r.cel[anoOrd]||{}).q)||0;
arr.sort((a,b)=>paisSort.col==="pais"?paisLabel(a.pais).localeCompare(paisLabel(b.pais),loc())*paisSort.dir:(qtdAno(a)-qtdAno(b))*paisSort.dir);
let grp='<tr class="grp"><th rowspan="2" class="stick" onclick="ordPais(\'pais\')" style="text-align:left">'+t("th_country")+(paisSort.col==="pais"?(paisSort.dir>0?" ▲":" ▼"):"")+'</th>';anos.forEach(a=>{const on=(paisSort.col==="qty"&&anoOrd===a);grp+='<th colspan="'+nm+'" onclick="ordPaisAno('+a+')" title="Ordenar por quantidade '+a+'">'+a+(on?(paisSort.dir>0?" ▲":" ▼"):"")+'</th>';});grp+='<th rowspan="2" class="var">'+t("th_var")+'</th></tr>';
let sub='<tr class="sub">';anos.forEach(()=>{mets.forEach((m,idx)=>{sub+='<th'+(idx===0?' class="b"':'')+'>'+t(m.i)+'</th>';});});sub+='</tr>';tbPaisHead.innerHTML=grp+sub;
tbPaisBody.innerHTML=arr.map(r=>{let tds='<td class="nm stick" onclick="selPais(\''+r.pais.replace(/'/g,"\\'")+'\')">'+paisLabel(r.pais)+'</td>';anos.forEach(a=>{const c=r.cel[a]||{q:0,f:0,c:0};mets.forEach((m,idx)=>{const val=celVal(c,m);const has=c.q||c.f;tds+='<td'+(idx===0?' class="b"':'')+'>'+(has?fmtMet(m,val):"—")+'</td>';});});tds+='<td>'+varCell(r.v1,r.v0)+'</td>';return'<tr class="'+(paisSel===r.pais?"sel":"")+'" onclick="selPais(\''+r.pais.replace(/'/g,"\\'")+'\')">'+tds+'</tr>';}).join("");
let ttl='<td class="nm stick" style="font-weight:800">'+t("total")+'</td>';anos.forEach(a=>{const dd=FILTRADO.filter(d=>d.ano===a);const c={q:soma(dd,"quantidade_un"),f:soma(dd,"valor_fob_usd"),c:soma(dd,"valor_cif_usd")};mets.forEach((m,idx)=>{ttl+='<td'+(idx===0?' class="b"':'')+' style="font-weight:800">'+fmtMet(m,celVal(c,m))+'</td>';});});
const gv1=FILTRADO.filter(d=>d.ano===ultimo&&d.mes<=maxMes).reduce((s,d)=>s+(d[kVar]||0),0),gv0=anterior?FILTRADO.filter(d=>d.ano===anterior&&d.mes<=maxMes).reduce((s,d)=>s+(d[kVar]||0),0):0;ttl+='<td style="font-weight:800">'+varCell(gv1,gv0)+'</td>';
tbPaisBody.innerHTML+='<tr style="border-top:2px solid var(--red);background:#F5F6F8">'+ttl+'</tr>';
varHint.textContent=anterior?t("var_txt").replace("{met}",t(metVar.i)).replace("{y1}",ultimo).replace("{y0}",anterior).replace("{m}",mesNome(maxMes)):"";
renderPaisSeg();}
function varCell(v1,v0){if(!v0)return '<span class="var-0">—</span>';const p=(v1-v0)/v0*100;const cls=p>0.05?"var-up":(p<-0.05?"var-dn":"var-0");const ar=p>0.05?"▲":(p<-0.05?"▼":"■");return '<span class="'+cls+'">'+ar+" "+(p>=0?"+":"")+p.toFixed(1)+"%</span>";}
function selPais(p){paisSel=(paisSel===p)?null:p;renderPais();}
function ordPais(c){if(paisSort.col===c){paisSort.dir*=-1;}else{paisSort.col=c;paisSort.dir=(c==="pais"?1:-1);}renderPais();}
function ordPaisAno(ano){if(paisSort.col==="qty"&&paisSortAno===ano){paisSort.dir*=-1;}else{paisSort.col="qty";paisSortAno=ano;paisSort.dir=-1;}renderPais();}
function renderPaisSeg(){const base=paisSel?FILTRADO.filter(d=>d.pais===paisSel):FILTRADO;const met=metRosca();
const ps=agr(base,"segmento_hp",met.k),ss=ORDEM_SEG.filter(s=>ps[s]),tot=ss.reduce((s,x)=>s+ps[x],0);
if(charts.paisSeg)charts.paisSeg.destroy();
charts.paisSeg=new Chart(chPaisSeg,{type:"doughnut",data:{labels:ss.map(segLabel),datasets:[{data:ss.map(s=>ps[s]),backgroundColor:ss.map(s=>CORES_SEG[s]),borderWidth:3,borderColor:"#fff"}]},options:doOpts(met)});
segStats.innerHTML=ss.map(s=>{const v=ps[s],pc=tot?(v/tot*100):0;return '<div class="row"><span class="nm"><span class="dot" style="background:'+CORES_SEG[s]+'"></span>'+segLabel(s)+'</span><span><span class="vl">'+fmtMetC(met,v)+'</span> <span class="pc">'+pc.toFixed(1)+'%</span></span></div>';}).join("");
const nome=paisSel?paisLabel(paisSel):t("all_countries");segTitle.textContent=t("c_seg_of")+" "+nome;
segChip.innerHTML=paisSel?'<span class="chip" onclick="selPais(\''+paisSel.replace(/'/g,"\\'")+'\')">'+paisLabel(paisSel)+' <span>✕</span></span>':"";
// Rodapé: se métrica for FOB (padrão) mostra preço médio; senão mostra o total da métrica
const vv=varSel();
if(vv==="all"||vv==="valor_fob_usd"||vv==="preco_medio_fob_un"){const q=soma(base,"quantidade_un"),f=soma(base,"valor_fob_usd");segAvg.textContent=t("avg_txt").replace("{c}",nome).replace("{v}",fUsd(q?f/q:0));}
else{const totV=soma(base,met.k);segAvg.textContent=t("tot_txt").replace("{lbl}",t(met.i)).replace("{c}",nome).replace("{v}",fmtMetC(met,totV));}
renderSegFilters(met);}
function renderSegFilters(met){
// Resume os filtros ATIVOS que impactam esta rosca: país, anos, meses acumulados, variável
const anos=[...new Set(FILTRADO.map(d=>d.ano))].filter(v=>!isNaN(v)).sort((a,b)=>a-b);
const meses=[...new Set(FILTRADO.map(d=>d.mes))].filter(v=>!isNaN(v)).sort((a,b)=>a-b);
const paisTxt=paisSel?paisLabel(paisSel):(mSel("fPais").length?mSel("fPais").map(paisLabel).join(", "):t("all_countries"));
const anoTxt=anos.length?(anos.length===1?anos[0]:anos[0]+" – "+anos[anos.length-1]):"—";
// Meses: se forem contíguos a partir de Jan, mostra "Jan → X" (acumulado); senão lista os selecionados
let mesTxt;
const contig=meses.length&&meses[0]===1&&meses.every((m,i)=>m===i+1);
if(!meses.length)mesTxt="—";
else if(meses.length===12)mesTxt=t("m_all");
else if(contig)mesTxt="Jan → "+mesNome(meses[meses.length-1]);
else mesTxt=meses.map(mesNome).join(", ");
const chips=[
 ["📍",t("fb_country"),paisTxt],
 ["📅",t("fb_year"),anoTxt],
 ["🗓️",t("fb_month"),mesTxt],
 ["📊",t("fb_var"),t(met.i)],
];
segFilters.innerHTML=chips.map(c=>'<span class="fchip"><span class="ic">'+c[0]+'</span>'+c[1]+': <b>'+c[2]+'</b></span>').join("");}
function renderTabela(){const CV=colsVisiveis();thead.innerHTML=CV.map(c=>{const ar=sortState.col===c.k?(sortState.dir>0?" ▲":" ▼"):"";return'<th class="'+(c.num?'num':'')+'" onclick="ordenar(\''+c.k+'\')">'+t(c.i)+ar+'</th>';}).join("");
const dd=[...FILTRADO].sort((a,b)=>{let va=a[sortState.col],vb=b[sortState.col];if(typeof va==="number")return(va-vb)*sortState.dir;return String(va).localeCompare(String(vb),loc())*sortState.dir;});cnt.textContent="("+dd.length.toLocaleString(loc())+" "+t("lines")+")";const MAX=500;
tbody.innerHTML=dd.slice(0,MAX).map(d=>"<tr>"+CV.map(c=>{let v=d[c.k];if(c.tag){return'<td><span class="seg-tag" style="background:'+(CORES_SEG[v]||"#999")+'">'+segLabel(v)+'</span></td>';}if(c.num){return'<td class="num">'+(c.fmt==="usd"?fUsd(v):fInt(v))+'</td>';}if(c.tr==="pais")return'<td>'+paisLabel(v)+'</td>';if(c.tr==="desc")return'<td>'+descLabel(v)+'</td>';if(c.ctr)return'<td class="ctr">'+(v==null?"":v)+'</td>';return"<td>"+(v==null?"":v)+"</td>";}).join("")+"</tr>").join("");
if(dd.length>MAX)tbody.innerHTML+='<tr><td colspan="'+CV.length+'" style="text-align:center;color:#999;padding:14px">… '+t("rowsmore").replace("{a}",MAX).replace("{b}",dd.length.toLocaleString(loc()))+'</td></tr>';
tfoot.innerHTML="<tr><td>"+t("total")+"</td>"+CV.slice(1).map(c=>{if(c.k==="quantidade_un")return'<td class="num">'+fInt(soma(dd,"quantidade_un"))+"</td>";if(c.k==="valor_fob_usd")return'<td class="num">'+fUsd(soma(dd,"valor_fob_usd"))+"</td>";if(c.k==="valor_cif_usd")return'<td class="num">'+fUsd(soma(dd,"valor_cif_usd"))+"</td>";if(c.k==="preco_medio_fob_un"){const q=soma(dd,"quantidade_un"),f=soma(dd,"valor_fob_usd");return'<td class="num">'+fUsd(q?f/q:0)+"</td>";}return"<td></td>";}).join("")+"</tr>";}
function ordenar(c){if(sortState.col===c)sortState.dir*=-1;else{sortState.col=c;sortState.dir=1;}renderTabela();}
function doOpts(met){return{responsive:true,maintainAspectRatio:false,cutout:"58%",plugins:{legend:{position:"bottom",labels:{font:{size:11,family:"Inter"},padding:11,usePointStyle:true}},tooltip:{callbacks:{label:c=>" "+c.label+": "+fmtMetC(met||MET_DEF[1],c.raw)}}}};}
function exportarCSV(){const CV=colsVisiveis();const h=CV.map(c=>t(c.i)).join(";");const l=FILTRADO.map(d=>CV.map(c=>{let v=c.tag?segLabel(d[c.k]):(c.tr==="pais"?paisLabel(d[c.k]):(c.tr==="desc"?descLabel(d[c.k]):d[c.k]));if(typeof v==="number")v=Math.round(v);return'"'+String(v==null?"":v).replace(/"/g,'""')+'"';}).join(";"));const csv="\uFEFF"+[h].concat(l).join("\n"),b=new Blob([csv],{type:"text/csv;charset=utf-8"}),a=document.createElement("a");a.href=URL.createObjectURL(b);a.download="comexstat_tratores.csv";a.click();}
</script></body></html>
"""

if __name__=="__main__":
    PERIODO_DE=ANO_INICIO; PERIODO_ATE=f"{datetime.now().year}-12"
    ULTIMO_PUBLICADO=obter_ultimo_periodo()
    print(f"[info] Solicitando {PERIODO_DE} a {PERIODO_ATE} (último publicado: {ULTIMO_PUBLICADO}).")
    df=consultar_importacao_tratores(PERIODO_DE,PERIODO_ATE)
    pasta=resolver_pasta_destino(); carimbo=datetime.now().strftime("%Y%m%d_%H%M")
    destino=pasta/f"importacao_tratores_HP_{PERIODO_DE}_a_{ULTIMO_PUBLICADO}_{carimbo}.xlsx"
    exportar_excel(df,destino)
    gerar_html_do_excel(destino,pasta/"index.html",PERIODO_DE,ULTIMO_PUBLICADO)
    print(f"\n✅ Pronto! Dados de {PERIODO_DE} a {ULTIMO_PUBLICADO}.")
    print(f"   Excel:     {destino}\n   Dashboard: {pasta/'index.html'}")
    if not df.empty: print("\nPrévia:"); print(df.head(12).to_string(index=False))
